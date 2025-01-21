import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from calendar import monthrange
from datetime import datetime

st.set_page_config(page_title="HRMS Attendance Report", page_icon="🕒")

def check_punch_status(row):
    """
    Check punch in/out status and return appropriate status message
    """
    punch_in = pd.isna(row['Punch_In_Time'])
    punch_out = pd.isna(row['Punch_Out_Time'])
    
    if punch_in and punch_out:
        return 'AT'
    elif punch_in and not punch_out:
        return 'Morning Punch Miss'
    elif not punch_in and punch_out:
        return 'Evening Punch Miss'
    return None

def process_attendance(attendance_data, hrms_data):
    # Convert Punch Date to datetime and extract components
    attendance_data['Punch_Date'] = pd.to_datetime(attendance_data['Punch_Date'], errors='coerce')
    
    # Process punch times - handling the new format
    attendance_data['Punch_In_Time'] = pd.to_datetime(attendance_data['Punch_In_Time'], errors='coerce')
    attendance_data['Punch_Out_Time'] = pd.to_datetime(attendance_data['Punch_Out_Time'], errors='coerce')
    
    # Extract time components
    attendance_data['Time IN HH:MM'] = attendance_data['Punch_In_Time'].dt.strftime('%H:%M')
    attendance_data['Time OUT HH:MM'] = attendance_data['Punch_Out_Time'].dt.strftime('%H:%M')
    
    # Determine the month and year from the data
    first_date = attendance_data['Punch_Date'].min()
    if pd.isna(first_date):
        st.error("No valid dates found in attendance data")
        return None
    
    month = first_date.month
    year = first_date.year
    
    # Get the number of days in the month
    _, days_in_month = monthrange(year, month)

    # Output DataFrame setup
    output_columns = ['Employee Id', 'Employee Name', 'Late Count'] + \
                    [f'Day {day}' for day in range(1, days_in_month + 1)] + \
                    ['Leaves Count', 'PL Count', 'CL Count', 'LL Count', 'LWP Count']
    output_data = pd.DataFrame(columns=output_columns)

    # Process each employee
    for _, emp_hrms_row in hrms_data.iterrows():
        emp_id = emp_hrms_row['Employee Id']
        emp_name = emp_hrms_row['Employee Name']
        late_count = 0
        pl_count = cl_count = ll_count = lwp_count = 0

        emp_row = {'Employee Id': emp_id, 'Employee Name': emp_name, 'Late Count': 0}

        for day in range(1, days_in_month + 1):
            day_str = f'{day:02d}-{month:02d}-{year}'
            day_column = f'Day {day}'
            emp_row[day_column] = None

            if day_str in hrms_data.columns:
                hrms_value = emp_hrms_row[day_str]

                if hrms_value in ['HD', 'WOff']:
                    emp_row[day_column] = hrms_value
                    continue

                if hrms_value == 'Not Enrolled':
                    emp_row[day_column] = 'Not Enrolled'
                    continue
                
                if hrms_value in ['PL/PT', 'CL/PT']:
                    # Get punch records for the employee and day
                    punch_day_records = attendance_data[
                        (attendance_data['Employee_ID'] == emp_id) &
                        (attendance_data['Punch_Date'].dt.day == day) &
                        (attendance_data['Punch_Date'].dt.month == month) &
                        (attendance_data['Punch_Date'].dt.year == year)
                    ]

                    if punch_day_records.empty:
                        emp_row[day_column] = 'AT'
                    else:
                        punch_status = check_punch_status(punch_day_records.iloc[0])
                        if punch_status == 'AT':
                            emp_row[day_column] = 'AT'
                        else:
                            emp_row[day_column] = 'Half Day Leave'
                    continue

                if hrms_value in ['PL', 'CL', 'LL', 'LWP']:
                    emp_row[day_column] = hrms_value
                    if hrms_value == 'PL':
                        pl_count += 1
                    elif hrms_value == 'CL':
                        cl_count += 1
                    elif hrms_value == 'LL':
                        ll_count += 1
                    elif hrms_value == 'LWP':
                        lwp_count += 1
                    continue

                # Get all punch records for the day using the new date format
                punch_day_records = attendance_data[ 
                    (attendance_data['Employee_ID'] == emp_id) & 
                    (attendance_data['Punch_Date'].dt.day == day) & 
                    (attendance_data['Punch_Date'].dt.month == month) & 
                    (attendance_data['Punch_Date'].dt.year == year) 
                ]
                
                if hrms_value == 'PT':
                    if punch_day_records.empty:
                        emp_row[day_column] = 'AT'
                    else:
                        # Check punch status for each record
                        punch_status = check_punch_status(punch_day_records.iloc[0])
                        if punch_status:
                            emp_row[day_column] = punch_status
                        else:
                            punch_in_time = punch_day_records.iloc[0]['Time IN HH:MM']
                            shift_name = punch_day_records.iloc[0]['Shift_Name']

                            if shift_name.strip().lower() == 'general' and punch_in_time > '09:45':
                                emp_row[day_column] = f'GSL {punch_in_time}'  # Changed from 'General Shift Late'
                                late_count += 1
                            elif shift_name.strip().lower() == 'evening shift' and punch_in_time > '16:30':
                                emp_row[day_column] = f'ESL {punch_in_time}'  # Changed from 'Evening Shift Late'
                                late_count += 1
                            else:
                                emp_row[day_column] = 'PT'
                elif hrms_value == 'WFH':
                    emp_row[day_column] = 'WFH'

        emp_row['Late Count'] = late_count
        emp_row['Leaves Count'] = pl_count + cl_count + ll_count + lwp_count
        emp_row['PL Count'] = pl_count
        emp_row['CL Count'] = cl_count
        emp_row['LL Count'] = ll_count
        emp_row['LWP Count'] = lwp_count

        output_data = pd.concat([output_data, pd.DataFrame([emp_row])], ignore_index=True)

    # Generate Excel output
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_data.to_excel(writer, index=False, sheet_name='Attendance Report')
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']

        # Updated color mapping to include the new formats
        category_colors = {
            'HD': PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid'),
            'WOff': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
            'PL': PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid'),
            'CL': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),
            'LL': PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid'),
            'LWP': PatternFill(start_color='eb9c42', end_color='eb9c42', fill_type='solid'),
            'WFH': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),
            'Morning Punch Miss': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
            'Evening Punch Miss': PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid'),
            'AT': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'PT': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'Not Enrolled': PatternFill(start_color='eb4d4d', end_color='eb4d4d', fill_type='solid'),
            'Half Day': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        }

        # Apply green color to Employee Id and Employee Name columns
        pt_fill = category_colors['PT']
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.fill = pt_fill

        # Modified color application to handle the new GSL and ESL formats
        gsl_fill = PatternFill(start_color='d8aaf2', end_color='d8aaf2', fill_type='solid')
        esl_fill = PatternFill(start_color='83f7f0', end_color='83f7f0', fill_type='solid')
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=worksheet.max_column):
            for cell in row:
                if cell.value:
                    if cell.value.startswith('GSL'):
                        cell.fill = gsl_fill
                    elif cell.value.startswith('ESL'):
                        cell.fill = esl_fill
                    elif cell.value in category_colors:
                        cell.fill = category_colors[cell.value]

    output.seek(0)
    return output

# Streamlit Interface
st.title("Monthly Attendance Processing System!")

st.subheader("Upload Files")
attendance_file = st.file_uploader("Upload Biometric Data (Excel)", type=['xlsx'])
hrms_file = st.file_uploader("Upload HRMS Data (CSV)", type=['csv'])

if st.button("Process Files"):
    if attendance_file and hrms_file:
        try:
            attendance_data = pd.read_excel(attendance_file)
            hrms_data = pd.read_csv(hrms_file)

            output = process_attendance(attendance_data, hrms_data)
            
            if output:
                st.success("Processing complete! Download your file below.")
                st.download_button(
                    "Download Report",
                    data=output,
                    file_name="attendance_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"An error occurred while processing the files: {str(e)}")
    else:
        st.error("Please upload both files to proceed.")
