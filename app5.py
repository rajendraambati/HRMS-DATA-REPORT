import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

st.set_page_config(page_title="HRMS Attendance Report", page_icon="🕒")

def process_attendance(attendance_data, hrms_data):
    # Process punch-in times in attendance data
    attendance_data['Punch IN Time'] = pd.to_datetime(attendance_data['Punch IN Time'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
    attendance_data['Time HH:MM'] = attendance_data['Punch IN Time'].dt.strftime('%H:%M')
    attendance_data['Date'] = attendance_data['Punch IN Time'].dt.date

    # Output DataFrame setup
    output_columns = ['Employee Id', 'Employee Name', 'Late Count'] + [f'Day {day}' for day in range(1, 32)] + ['Leaves Count', 'PL Count', 'CL Count', 'LL Count', 'LWP Count']
    output_data = pd.DataFrame(columns=output_columns)

    # Process each employee
    for _, emp_hrms_row in hrms_data.iterrows():
        emp_id = emp_hrms_row['Employee Id']
        emp_name = emp_hrms_row['Employee Name']
        late_count = 0
        pl_count = cl_count = ll_count = lwp_count = 0

        emp_row = {'Employee Id': emp_id, 'Employee Name': emp_name, 'Late Count': 0}

        for day in range(1, 32):
            day_str = f'{day:02d}-01-2024'
            day_column = f'Day {day}'
            emp_row[day_column] = None

            if day_str in hrms_data.columns:
                hrms_value = emp_hrms_row[day_str]

                if hrms_value == 'Not Enrolled':
                    emp_row[day_column] = 'Not Enrolled'
                    continue

                if hrms_value == 'PL/PT':  # Check for 'PL/PT' or similar values
                    emp_row[day_column] = 'Half Day'
                    continue

                if hrms_value in ['HD', 'WOff']:
                    emp_row[day_column] = hrms_value
                    continue

                if hrms_value in ['PL', 'CL', 'LL', 'LWP']:
                    emp_row[day_column] = hrms_value
                    if hrms_value == 'PL':
                        pl_count += 1
                    elif hrms_value == 'CL':
                        cl_count += 1
                    elif hrms_value == 'LL':
                        ll_count += 1
                    elif hrms_value == 'LWP':
                        lwp_count += 1
                    continue

                punch_day_records = attendance_data[
                    (attendance_data['employee_id'] == emp_id) & 
                    (attendance_data['Punch IN Time'].dt.day == day)
                ]

                if hrms_value == 'PT':
                    if not punch_day_records.empty:
                        punch_in_time = punch_day_records.iloc[0]['Time HH:MM']
                        shift_name = punch_day_records.iloc[0]['shift_name']

                        if shift_name.strip().lower() == 'general' and punch_in_time > '09:45':
                            emp_row[day_column] = 'General Shift Late'
                            late_count += 1
                        elif shift_name.strip().lower() == 'evening shift' and punch_in_time > '14:30':
                            emp_row[day_column] = 'Evening Shift Late'
                            late_count += 1
                        else:
                            emp_row[day_column] = 'PT'
                    else:
                        emp_row[day_column] = 'Punch Miss'
                elif hrms_value == 'WFH':
                    emp_row[day_column] = 'WFH'

        emp_row['Late Count'] = late_count
        emp_row['Leaves Count'] = pl_count + cl_count + ll_count + lwp_count
        emp_row['PL Count'] = pl_count
        emp_row['CL Count'] = cl_count
        emp_row['LL Count'] = ll_count
        emp_row['LWP Count'] = lwp_count

        output_data = pd.concat([output_data, pd.DataFrame([emp_row])], ignore_index=True)

    # Generate Excel output
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_data.to_excel(writer, index=False, sheet_name='Attendance Report')
        workbook = writer.book
        worksheet = writer.sheets['Attendance Report']

        # Formatting: Define fills for each category
        category_colors = {
            'HD': PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid'),  # Light Steel Blue
            'WOff': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),  # Light Gray
            'PL': PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid'),  # Pale Green
            'CL': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),  # Light Blue
            'LL': PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid'),  # Light Salmon
            'LWP': PatternFill(start_color='FF4500', end_color='FF4500', fill_type='solid'),  # Orange Red
            'WFH': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),  # Lemon Chiffon
            'General Shift Late': PatternFill(start_color='d8aaf2', end_color='d8aaf2', fill_type='solid'),  # Gold
            'Evening Shift Late': PatternFill(start_color='83f7f0', end_color='83f7f0', fill_type='solid'),  # Dark Orange
            'Punch Miss': PatternFill(start_color='6969f0', end_color='6969f0', fill_type='solid'),  # Tomato
            'PT': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),  # Green
            'Not Enrolled': PatternFill(start_color='eb4d4d', end_color='eb4d4d', fill_type='solid'),  # Dark Red
            'Half Day': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
        }

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=worksheet.max_column):
            for cell in row:
                if cell.value in category_colors:
                    cell.fill = category_colors[cell.value]

    output.seek(0)
    return output

# Streamlit Interface
st.title("Monthly Attendance Processing System!")

st.subheader("Upload Files")
attendance_file = st.file_uploader("Upload Biometric Data (Excel)", type=['xlsx'])
hrms_file = st.file_uploader("Upload HRMS Data (CSV)", type=['csv'])

if st.button("Process Files"):
    if attendance_file and hrms_file:
        attendance_data = pd.read_excel(attendance_file)
        hrms_data = pd.read_csv(hrms_file)

        output = process_attendance(attendance_data, hrms_data)

        st.success("Processing complete! Download your file below.")
        st.download_button("Download Report", data=output, file_name="attendance_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.error("Please upload both files to proceed.")
