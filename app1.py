import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Step 1: Load datasets
attendance_data = pd.read_excel('b.xlsx')
hrms_data = pd.read_csv('km.csv')

# Process punch-in times in attendance data
attendance_data['Punch IN Time'] = pd.to_datetime(attendance_data['Punch IN Time'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
attendance_data['Time HH:MM'] = attendance_data['Punch IN Time'].dt.strftime('%H:%M')
attendance_data['Date'] = attendance_data['Punch IN Time'].dt.date

# Output DataFrame setup
output_columns = ['Employee Id', 'Employee Name', 'Late Count'] + [f'Day {day}' for day in range(1, 32)] + ['Leaves Count', 'PL Count', 'CL Count', 'LL Count', 'LWP Count']
output_data = pd.DataFrame(columns=output_columns)

# Process each employee
for _, emp_hrms_row in hrms_data.iterrows():
    emp_id = emp_hrms_row['Employee Id']
    emp_name = emp_hrms_row['Employee Name']
    late_count = 0
    pl_count = cl_count = ll_count = lwp_count = 0

    # Row initialization
    emp_row = {'Employee Id': emp_id, 'Employee Name': emp_name, 'Late Count': 0}

    for day in range(1, 32):
        day_str = f'{day:02d}-01-2024'  # Create date string
        day_column = f'Day {day}'
        emp_row[day_column] = None

        # If the column exists in HRMS data
        if day_str in hrms_data.columns:
            hrms_value = emp_hrms_row[day_str]

            if hrms_value in ['HD', 'WOff']:
                emp_row[day_column] = hrms_value
                continue

            if hrms_value in ['PL', 'CL', 'LL', 'LWP']:
                emp_row[day_column] = hrms_value
                if hrms_value == 'PL':
                    pl_count += 1
                elif hrms_value == 'CL':
                    cl_count += 1
                elif hrms_value == 'LL':
                    ll_count += 1
                elif hrms_value == 'LWP':
                    lwp_count += 1
                continue

            # Process attendance based on punch-in data
            punch_day_records = attendance_data[(
                attendance_data['employee_id'] == emp_id) & 
                (attendance_data['Punch IN Time'].dt.day == day)
            ]

            if hrms_value == 'PT':
                if not punch_day_records.empty:
                    punch_in_time = punch_day_records.iloc[0]['Time HH:MM']
                    shift_name = punch_day_records.iloc[0]['shift_name']

                    if shift_name.strip().lower() == 'general' and punch_in_time > '09:45':
                        emp_row[day_column] = 'General Shift Late'
                        late_count += 1

                    elif shift_name.strip().lower() == 'evening shift' and punch_in_time > '14:30':
                        emp_row[day_column] = 'Evening Shift Late'
                        late_count += 1
                    else:
                        emp_row[day_column] = 'PT'
                else:
                    emp_row[day_column] = 'Punch Miss'
            elif hrms_value == 'WFH':
                emp_row[day_column] = 'WFH'

    emp_row['Late Count'] = late_count
    emp_row['Leaves Count'] = pl_count + cl_count + ll_count + lwp_count
    emp_row['PL Count'] = pl_count
    emp_row['CL Count'] = cl_count
    emp_row['LL Count'] = ll_count
    emp_row['LWP Count'] = lwp_count

    output_data = pd.concat([output_data, pd.DataFrame([emp_row])], ignore_index=True)

# Step 5: Export to Excel with formatting
excel_file_path = 'employee_attendance_report.xlsx'
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    output_data.to_excel(writer, index=False, sheet_name='Attendance Report')

    workbook = writer.book
    worksheet = writer.sheets['Attendance Report']

    # Define fill styles
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    light_green_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')
    purple_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Apply formatting
    for row in range(2, len(output_data) + 2):
        worksheet[f'A{row}'].fill = green_fill
        worksheet[f'B{row}'].fill = green_fill

        for day in range(1, 32):
            col_letter = get_column_letter(day + 3)
            cell = worksheet[f'{col_letter}{row}']

            if cell.value:
                if cell.value == 'HD':
                    cell.fill = blue_fill
                elif cell.value == 'WOff':
                    cell.fill = yellow_fill
                elif cell.value == 'PT':
                    cell.fill = light_green_fill
                elif cell.value == 'Punch Miss':
                    cell.fill = red_fill
                elif ':' in str(cell.value):  # Late time
                    cell.fill = red_fill
                elif cell.value in ['PL', 'CL', 'LWP', 'LL']:
                    cell.fill = purple_fill

    # Adjust column widths
    for column in worksheet.columns:
        max_length = max(len(str(cell.value) or "") for cell in column)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

print("Attendance report generated successfully!")
